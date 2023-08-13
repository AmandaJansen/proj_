import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import matplotlib.pyplot as plt
import pyrebase
import subprocess #Responsável por executar outro arquivo "Py"

#Importando biblioteca para utilizar na SDK

import firebase_admin
from firebase_admin import credentials


config = {
  "apiKey": "AIzaSyC9-RdrYzOB3tfuxaiQDS1RxlLQAnojFgI",
  "authDomain": "python-1c86a.firebaseapp.com",
  "databaseURL": "https://python-1c86a-default-rtdb.firebaseio.com",
  "pojectId": "python-1c86a",
  "storageBucket": "python-1c86a.appspot.com",
  "messagingSenderId": "48968104594",
  "appId": "1:48968104594:web:e24e04b38f44cb26ae044b"
}


firebase = pyrebase.initialize_app(config)
db = firebase.database()

cred = credentials.Certificate("sdk.json")
firebase_admin.initialize_app(cred, {'databaseURL': 'https://bdpython-5fd52-default-rtdb.firebaseio.com'})


# Função para adicionar o cliente no array
def adicionar_cliente():
    nome = entry_nome.get() # pegar/buscar o que o usuário digitou no campo detexto
    data_nascimento = entry_data_nascimento.get()
    cpf = entry_cpf.get()
    origem = entry_origem.get()
    score = int(entry_score.get())
    usuario = (entry_usuario.get())
    senha = (entry_senha.get())
   

    data = {
        'Nome': nome,
        'Data de Nascimento': data_nascimento,
        'CPF': cpf,
        'Origem': origem,
        'Score': score,
        'Usuario':usuario,
        'Senha': senha
    }
    db.child("clientes").push(data)


    # Limpar os campos após adicionar o cliente
    entry_nome.delete(0, tk.END)
    entry_data_nascimento.delete(0, tk.END)
    entry_cpf.delete(0, tk.END)
    entry_origem.delete(0, tk.END)
    entry_score.delete(0, tk.END)
    entry_usuario.delete(0, tk.END)
    entry_senha.delete(0, tk.END)


    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")


def salvar_em_excel():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score'])


    # Adiciona os dados dos clientes na planilha a partir da última linha disponível
    for cliente in clientes:
        sheet.append([cliente['Nome'], cliente['Data de Nascimento'], cliente['CPF'], cliente['Origem'], cliente['Score']])


    # Salva a planilha em um arquivo
    workbook.save('clientes.xlsx')


    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")


def consultar_dados():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        nome_consulta = entry_consulta_nome.get().lower()  # Converte o nome digitado para minúsculo
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():  # Verifica se o nome digitado está contido no nome do cliente cadastrado
                dados.append(f"Nome: {row[0]}, Data de Nascimento: {row[1]}, CPF: {row[2]}, Origem: {row[3]}, Score: {row[4]}")
       
        if dados:
            messagebox.showinfo("Dados dos Clientes", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum cliente encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")


def criar_grafico():
    # Obtendo dados do Realtime Database
    ref = db.child('clientes')
    clientes_from_db = ref.get().val()
   
    #Transformando os dados em uma lista (caso ainda não estejam nesse formato)
    if isinstance(clientes_from_db, dict):
         clientes_from_db = list(clientes_from_db.values())
         
    
    opcao_selecionada = selected_option.get()
    
    if opcao_selecionada == 'score':
        valores_x = [cliente['Nome'] for cliente in clientes_from_db if isinstance(cliente, dict)]
        valores_y = [cliente['Score'] for cliente in clientes_from_db if isinstance(cliente, dict)]
        titulo = 'Scores dos Clientes'
        eixo_x = 'Clientes'
        eixo_y = 'Score'
        
    elif opcao_selecionada == "origem":
        valores_x = [cliente['Nome'] for cliente in clientes_from_db]
        valores_y = [cliente['Origem'] for cliente in clientes_from_db]
        titulo = 'Origem dos Clientes'
        eixo_x = 'Clientes'
        eixo_y = 'Origem'


    plt.bar(valores_x, valores_y)
    plt.xlabel(eixo_x)
    plt.ylabel(eixo_y)
    plt.title(titulo)
    plt.xticks(rotation=45)
    plt.show()
    
    
 #Variável global para armazenar o ID do cliente atualmente carrregado no formulário      
cliente_atual_id = None
        
def carregar_dados():
    global cliente_atual_id


    cpf = entry_cpf.get()
   
    # Obtendo o cliente pelo CPF
    cliente_data = db.child("clientes").order_by_child("CPF").equal_to(cpf).get().val()
   
    # Se o cliente for encontrado, carregue os detalhes nos campos de entrada
    if cliente_data:
        cliente_id, cliente_info = list(cliente_data.items())[0]
        cliente_atual_id = cliente_id
       
        entry_nome.insert(0, cliente_info['Nome'])
        entry_data_nascimento.insert(0, cliente_info['Data de Nascimento'])
        entry_origem.insert(0, cliente_info['Origem'])
        entry_score.insert(0, str(cliente_info['Score']))
        entry_usuario.insert(0, cliente_info['Usuario'])
        entry_senha.insert(0, cliente_info['Senha'])
                       
    button_alterar['state'] = 'normal'  # Habilitar o botão de alterar 
   
       
def atualizar_cliente():
    if cliente_atual_id:
        novo_data = {
            'Nome': entry_nome.get(),
            'Data de Nascimento': entry_data_nascimento.get(),
            'Origem': entry_origem.get(),
            'Score': entry_score.get(),
            'Usuario': entry_usuario.get(),
            'Senha': entry_senha.get()      
        }
        db.child("clientes").child(cliente_atual_id).update(novo_data)
        messagebox.showinfo("Sucesso", "Dados do cliente atualizados com sucesso!")
    else:
        messagebox.showerror("Erro", "Nenhum cliente carregado para atualizar!")

def abrir_tela_principal():
    window.destroy() #Fecha a janela atual antes de abrir a tela principal
    subprocess.run(["phython", "tela_main.py" ])
    

# Criação da janela
window = tk.Tk()
window.configure(bg="lightgreen")
window.title("Cadastro de Clientes")
window.geometry("400x690")


# Array para armazenar os clientes
clientes = []


# Labels e campos de entrada
label_nome = tk.Label(window, text="Nome:",bg="lightgreen")
label_nome.pack()


entry_nome = tk.Entry(window)
entry_nome.pack()


label_data_nascimento = tk.Label(window, text="Data de Nascimento:",bg="lightgreen")
label_data_nascimento.pack()


entry_data_nascimento = tk.Entry(window)
entry_data_nascimento.pack()


label_cpf = tk.Label(window, text="CPF:",bg="lightgreen")
label_cpf.pack()


entry_cpf = tk.Entry(window)
entry_cpf.pack()


label_origem = tk.Label(window, text="Origem (Loja ou Site):",bg="lightgreen")
label_origem.pack()


entry_origem = tk.Entry(window)
entry_origem.pack()


label_score = tk.Label(window, text="Score:",bg="lightgreen")
label_score.pack()


entry_score = tk.Entry(window)
entry_score.pack()

label_usuario = tk.Label(window, text="Usuario:",bg="lightgreen")
label_usuario.pack()


entry_usuario = tk.Entry(window)
entry_usuario.pack()

label_senha = tk.Label(window, text="Senha:",bg="lightgreen")
label_senha.pack()


entry_senha = tk.Entry(window)
entry_senha.pack()


# Botão para adicionar cliente
button_adicionar = tk.Button(window, text="Adicionar Cliente", command=adicionar_cliente)
button_adicionar.pack(pady=8)


# Botão para salvar os dados em Excel
button_salvar_excel = tk.Button(window, text="Salvar em Excel", command=salvar_em_excel)
button_salvar_excel.pack(pady=8)


# Campo de consulta por nome
label_consulta_nome = tk.Label(window, text="Consultar por Nome:",bg="lightgreen")
label_consulta_nome.pack(pady=8)


entry_consulta_nome = tk.Entry(window)
entry_consulta_nome.pack(pady=8)


# Botão para consultar os dados
button_consultar_dados = tk.Button(window, text="Consultar Dados", command=consultar_dados)
button_consultar_dados.pack()

# Variável para armazenar a escolha do usuário
selected_option = tk.StringVar()


label_opcoes = tk.Label(window, text="Escolha os dados para o relatório:",bg="lightgreen")
label_opcoes.pack()


radio_score = tk.Radiobutton(window, text="Score", variable=selected_option, value="score",bg="lightgreen")
radio_score.pack()


radio_origem = tk.Radiobutton(window, text="Origem", variable=selected_option, value="origem",bg="lightgreen")
radio_origem.pack()


# Botão para criar o gráfico
button_criar_grafico = tk.Button(window, text="Criar Gráfico", command=criar_grafico)
button_criar_grafico.pack(pady=8)

# Botão para abrir a tela principal
button_voltar = tk.Button(window, text="Voltar para Tela Principal", command=abrir_tela_principal)
button_voltar.pack()


#Botão para carregar dados
button_carregar = tk.Button(window, text="Carregar Dados", command=carregar_dados)
button_carregar.pack()

#Botão para alterar dados
button_alterar = tk.Button(window, text="Alterar Dados", command=atualizar_cliente, state='disabled')  # Inicia desabilitado
button_alterar.pack()




window.mainloop()
