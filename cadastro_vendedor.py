import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

def cadastrar_vendedor():
    nome = entry_nome.get()
    matricula = float(entry_matricula.get())
    

    vendedores.append({"Nome": nome, "Matricula": matricula})
    entry_nome.delete(0, tk.END)
    entry_matricula.delete(0, tk.END)
   

def exibir_vendedores():
    resultado_text.delete(1.0, tk.END)
    for vendedor in vendedores:
        resultado_text.insert(tk.END, f"Nome: {vendedor['Nome']}\n")
        resultado_text.insert(tk.END, f"Matricula: R$ {vendedor['Matricula']:.2f}\n")
        resultado_text.insert(tk.END, "----------------\n")

vendedores = []

def salvar_vendedor_em_excel():
    try:
        workbook = load_workbook('vendedores.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Matricula'])

    # Adiciona os dados dos vendedor na planilha a partir da última linha disponível
    for vendedor in vendedores:
        sheet.append([vendedor['Nome'], vendedor['Matricula']])

    # Salva a planilha em um arquivo
    workbook.save('vendedor.xlsx')

    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")

window = tk.Tk()
window.title("Cadastro de Vendedores")
window.geometry("300x400")

label_nome = tk.Label(window, text="Nome:")
label_nome.pack()

entry_nome = tk.Entry(window)
entry_nome.pack()

label_matricula = tk.Label(window, text="Matricula:")
label_matricula.pack()

entry_matricula = tk.Entry(window)
entry_matricula.pack()

button_cadastrar = tk.Button(window, text="Cadastrar", command=cadastrar_vendedor)
button_cadastrar.pack()

button_exibir = tk.Button(window, text="Exibir Vendedores", command=exibir_vendedores)
button_exibir.pack()

resultado_text = tk.Text(window, height=10, width=30)
resultado_text.pack()

button_salvar_excel = tk.Button(window, text="Salvar Vendedores em Excel", command=salvar_vendedor_em_excel)
button_salvar_excel.pack()

window.mainloop()