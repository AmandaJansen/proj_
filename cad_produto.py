import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

def cadastrar_produto():
    nome = entry_nome.get()
    valor = float(entry_valor.get())
    categoria = entry_categoria.get()

    produtos.append({"Nome": nome, "Valor": valor, "Categoria": categoria})
    entry_nome.delete(0, tk.END)
    entry_valor.delete(0, tk.END)
    entry_categoria.delete(0, tk.END)

def exibir_produtos():
    resultado_text.delete(1.0, tk.END)
    for produto in produtos:
        resultado_text.insert(tk.END, f"Nome: {produto['Nome']}\n")
        resultado_text.insert(tk.END, f"Valor: R$ {produto['Valor']:.2f}\n")
        resultado_text.insert(tk.END, f"Categoria: {produto['Categoria']}\n")
        resultado_text.insert(tk.END, "----------------\n")

produtos = []

def salvar_produtos_em_excel():
    try:
        workbook = load_workbook('produtos.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Valor', 'Categoria'])

    # Adiciona os dados dos produtos na planilha a partir da última linha disponível
    for produto in produtos:
        sheet.append([produto['Nome'], produto['Valor'], produto['Categoria']])

    # Salva a planilha em um arquivo
    workbook.save('produtos.xlsx')

    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")

window = tk.Tk()
window.title("Cadastro de Produtos")
window.configure(bg="lightyellow")
window.geometry("300x500")

label_nome = tk.Label(window, text="Nome:", bg="lightyellow")
label_nome.pack()

entry_nome = tk.Entry(window)
entry_nome.pack()

label_valor = tk.Label(window, text="Valor:",bg="lightyellow")
label_valor.pack()

entry_valor = tk.Entry(window)
entry_valor.pack()

label_categoria = tk.Label(window, text="Categoria:",bg="lightyellow")
label_categoria.pack()

entry_categoria = tk.Entry(window)
entry_categoria.pack()

button_cadastrar = tk.Button(window, text="Cadastrar", command=cadastrar_produto)
button_cadastrar.pack(pady=8)

button_exibir = tk.Button(window, text="Exibir Produtos", command=exibir_produtos)
button_exibir.pack(pady=8)

resultado_text = tk.Text(window, height=10, width=30)
resultado_text.pack()

button_salvar_excel = tk.Button(window, text="Salvar Produtos em Excel", command=salvar_produtos_em_excel)
button_salvar_excel.pack(pady=8)

window.mainloop()