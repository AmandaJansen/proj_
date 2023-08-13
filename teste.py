import tkinter as tk
import subprocess
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import matplotlib.pyplot as plt
import traceback

# Função decoradora para capturar exceções e exibir traceback
def excecao_com_traceback(funcao):
    def wrapper(*args, **kwargs):
        try:
            return funcao(*args, **kwargs)
        except Exception as e:
            traceback_str = traceback.format_exc()
            print(traceback_str)
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}\n\nTraceback:\n{traceback_str}")
    return wrapper

# Função para adicionar o cliente no array
@excecao_com_traceback
def adicionar_cliente():
    nome = entry_nome.get()
    data_nascimento = entry_data_nascimento.get()
    cpf = entry_cpf.get()
    origem = entry_origem.get()
    score = int(entry_score.get())

    clientes.append({
        'Nome': nome,
        'Data de Nascimento': data_nascimento,
        'CPF': cpf,
        'Origem': origem,
        'Score': score
    })

    entry_nome.delete(0, tk.END)
    entry_data_nascimento.delete(0, tk.END)
    entry_cpf.delete(0, tk.END)
    entry_origem.delete(0, tk.END)
    entry_score.delete(0, tk.END)

    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")

# Função para salvar os dados em Excel
@excecao_com_traceback
def salvar_em_excel():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score'])

    for cliente in clientes:
        sheet.append([cliente['Nome'], cliente['Data de Nascimento'], cliente['CPF'], cliente['Origem'], cliente['Score']])

    workbook.save('clientes.xlsx')
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")

# Função para consultar os dados
@excecao_com_traceback
def consultar_dados():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        nome_consulta = entry_consulta_nome.get().lower()
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():
                dados.append(f"Nome: {row[0]}, Data de Nascimento: {row[1]}, CPF: {row[2]}, Origem: {row[3]}, Score: {row[4]}")
        
        if dados:
            messagebox.showinfo("Dados dos Clientes", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum cliente encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")

# Função para criar o gráfico
@excecao_com_traceback
def criar_grafico():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        clientes_from_excel = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = {
                'Nome': row[0],
                'Data de Nascimento': row[1],
                'CPF': row[2],
                'Origem': row[3],
                'Score': row[4]
            }
            clientes_from_excel.append(cliente)

        opcao_selecionada = selected_option.get()

        valores_x = []
        valores_y = []

        if opcao_selecionada == "score1":
            valores_x = [cliente['Nome'] for cliente in clientes_from_excel]
            valores_y = [cliente['Score'] for cliente in clientes_from_excel]
            titulo = 'Scores dos Clientes'
            eixo_x = 'Clientes'
            eixo_y = 'Score'
        elif opcao_selecionada == "origem":
            valores_x = [cliente['Nome'] for cliente in clientes_from_excel]
            valores_y = [cliente['Origem'] for cliente in clientes_from_excel]
            titulo = 'Origem dos Clientes'
            eixo_x = 'Clientes'
            eixo_y = 'Origem'
            
        elif opcao_selecionada == "score2":
            valores_x = [cliente['Nome'] for cliente in clientes_from_excel]
            valores_y = [cliente['Score'] for cliente in clientes_from_excel]
            titulo = 'Scores dos Clientes'
            eixo_x = 'Clientes'
            eixo_y = 'Score'
        
        else:
            messagebox.showerror("Erro", "Opção inválida selecionada para criar o gráfico.")
            return

        plt.bar(valores_x, valores_y)
        plt.xlabel(eixo_x)
        plt.ylabel(eixo_y)
        plt.title(titulo)
        plt.xticks(rotation=45)
        plt.show()
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")

# Função para abrir a tela principal
@excecao_com_traceback
def abrir_tela_principal():
    window.destroy()
    subprocess.run(["python", "telaPrincipal.py"])

# Criação da janela
window = tk.Tk()
window.title("Cadastro de Clientes")
window.geometry("400x550")

clientes = []

label_nome = tk.Label(window, text="Nome:")
label_nome.pack()

entry_nome = tk.Entry(window)
entry_nome.pack()

label_data_nascimento = tk.Label(window, text="Data de Nascimento:")
label_data_nascimento.pack()

entry_data_nascimento = tk.Entry(window)
entry_data_nascimento.pack()

label_cpf = tk.Label(window, text="CPF:")
label_cpf.pack()

entry_cpf = tk.Entry(window)
entry_cpf.pack()

label_origem = tk.Label(window, text="Origem (Loja ou Site):")
label_origem.pack()

entry_origem = tk.Entry(window)
entry_origem.pack()

label_score = tk.Label(window, text="Score:")
label_score.pack()

entry_score = tk.Entry(window)
entry_score.pack()

botao_voltar = tk.Button(window, text="Voltar para Tela Principal", command=abrir_tela_principal)
botao_voltar.pack()

# Campo de consulta por nome
label_consulta_nome = tk.Label(window, text="Consultar por Nome:")
label_consulta_nome.pack()

entry_consulta_nome = tk.Entry(window)
entry_consulta_nome.pack()

# Botão para consultar os dados
button_consultar_dados = tk.Button(window, text="Consultar Dados", command=consultar_dados)
button_consultar_dados.pack()

# Variável para armazenar a escolha do usuário
selected_option = tk.StringVar()

label_opcoes = tk.Label(window, text="Escolha os dados para o relatório:")
label_opcoes.pack()

radio_score = tk.Radiobutton(window, text="Score", variable=selected_option, value="score2")
radio_score.pack()

radio_origem = tk.Radiobutton(window, text="Origem", variable=selected_option, value="origem")
radio_origem.pack()

# Botão para criar o gráfico
button_criar_grafico = tk.Button(window, text="Criar Gráfico", command=criar_grafico)
button_criar_grafico.pack()

window.mainloop()

