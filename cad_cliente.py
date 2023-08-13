import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt


def addCliente():
  #criando uma funçaão onde pega  os dados inseridos pelo usuario no input
  name = entry_name.get() #cria um avariavel onde pega(get) o input(entry_name) para ler os dados inseridos
  data_nascimento = entry_nasc.get()
  cpf = entry_cpf.get()
  origem = entry_origem.get()
  score = entry_score.get()

  clientes.append({
        'Nome': name,
        'Data de Nascimento': data_nascimento,
        'CPF': cpf,
        'Origem': origem,
        'Score': score
    })
  
  # Limpar os campos após adicionar o cliente
  entry_name.delete(0, tk.END) #pega o input(entry_name) deleta os dados inseridos(.delete) da posição 0 até o fim apague tudo (0, tk.END)
  entry_nasc.delete(0, tk.END)
  entry_cpf.delete(0, tk.END)
  entry_origem.delete(0, tk.END)
  entry_score.delete(0, tk.END)

  # Exibir mensagem de sucesso
  messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")


def salvar_em_excel():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Data de Nascimento', 'CPF', 'Origem', 'Score'])

    # Adiciona os dados dos clientes na planilha a partir da última linha disponível
    for cliente in clientes:
        sheet.append([cliente['Nome'], cliente['Data de Nascimento'], cliente['CPF'], cliente['Origem'], cliente['Score']])

    # Salva a planilha em um arquivo
    workbook.save('clientes.xlsx')

    # Exibir mensagem de sucesso
    messagebox.showinfo("Sucesso", "Dados salvos no Excel!")


def consultar_dados():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        nome_consulta = entry_consulta_nome.get().lower()  # Converte o nome digitado para minúsculo
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if nome_consulta in row[0].lower():  # Verifica se o nome digitado está contido no nome do cliente cadastrado
                dados.append(f"Nome: {row[0]}, Data de Nascimento: {row[1]}, CPF: {row[2]}, Origem: {row[3]}, Score: {row[4]}")
        
        if dados:
            messagebox.showinfo("Dados dos Clientes", "\n".join(dados))
        else:
            messagebox.showinfo("Consulta", "Nenhum cliente encontrado com o nome informado.")
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")


def criar_grafico():
    try:
        workbook = load_workbook('clientes.xlsx')
        sheet = workbook.active
        clientes_from_excel =[]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = {
                'Nome': row[0],
                'Data de nascimento': row[1],
                'CPF':row[2],
                'Origem': row[3],
                'Score': row[4]
            }
            clientes_from_excel.append(cliente)
        nomes = [cliente['Nome'] for cliente in clientes_from_excel]
        scores =[cliente['Score'] for cliente in clientes_from_excel]
        
        plt.bar(nomes, scores)
        plt.xlabel('Clientes')
        plt.ylabel('Score')
        plt.title('Scores dos Cliente')
        plt.xticks(rotation=45)
        plt.show()
        
    except FileNotFoundError:
        messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")
    
    
# Criar a janela principal
myWindow = tk.Tk()                 #cria a tela 
myWindow.configure(bg="lightblue") #define a cor de fundo da tela
myWindow.title("Cliente")          #titulo da tela
myWindow.geometry("300x550")       #define o tamanho da tela

#Cria o array para armazenar os clientes
clientes = []

# Criar widgets
label_name = tk.Label(myWindow, text="Digite seu nome completo: ", font=("Arial", 10, "bold"), bg="lightblue",fg="black")
label_name.pack(pady=10)

entry_name = tk.Entry(myWindow) #tk.Entry esta puxando o input , entre os parentese esta falando a onde o input vai aparecer
entry_name.pack(pady=5) #o pack é para deixar visivel na tela 

label_nasc = tk.Label(myWindow, text="Insira a data de nascimento: ", font=("Arial", 10, "bold"), bg="lightblue",fg="black") #tk.Label esta puxando a label biblioteca tkinter
label_nasc.pack(pady=10)

entry_nasc = tk.Entry(myWindow)
entry_nasc.pack(pady=5)

label_cpf = tk.Label(myWindow, text="Digite o CPF: ", font=("Arial", 10, "bold"), bg="lightblue",fg="black")
label_cpf.pack(pady=10)

entry_cpf = tk.Entry(myWindow)
entry_cpf.pack(pady=5)

label_origem = tk.Label(myWindow, text="Insira a origem (Loja ou Site): ", font=("Arial", 10, "bold"), bg="lightblue",fg="black")
label_origem.pack(pady=10)

entry_origem = tk.Entry(myWindow)
entry_origem.pack(pady=5)

label_score = tk.Label(myWindow, text="Insira o Score: ", font=("Arial", 10, "bold"), bg="lightblue",fg="black")
label_score.pack(pady=10)

entry_score = tk.Entry(myWindow)
entry_score.pack(pady=5)


# Botão para adicionar cliente
button_adicionar = tk.Button(myWindow, text="Adicionar Cliente", font=("Arial", 9, "bold"), bg="gray",fg="white", command=addCliente)
button_adicionar.pack(pady=5)

# Botão para salvar os dados em Excel
button_salvar_excel = tk.Button(myWindow, text="Salvar em Excel",  font=("Arial", 9, "bold"), bg="gray",fg="white", command=salvar_em_excel)
button_salvar_excel.pack(pady=5)

# Campo de consulta por nome
label_consulta_nome = tk.Label(myWindow, text="Consultar por Nome:" , font=("Arial", 10, "bold"), bg="lightblue",fg="black")
label_consulta_nome.pack(pady=5)

entry_consulta_nome = tk.Entry(myWindow)
entry_consulta_nome.pack(pady=5)

# Botão para consultar os dados
button_consultar_dados = tk.Button(myWindow, text="Consultar Clientes", font=("Arial", 9, "bold"), bg="gray",fg="white", command=consultar_dados)
button_consultar_dados.pack(pady=5)


# Iniciar o loop principal da interface gráfica
myWindow.mainloop()




